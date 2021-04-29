"""
Module to convert excel files to txt files.
"""

import os
import sys
from typing import List, Union, Optional, NamedTuple
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Arguments(NamedTuple):
    """
    Result of the command line arguments.
    """
    inputPath: str
    sheetName: Optional[str]
    outputPath: Optional[str]

    def __str__(self) -> str:
        return f'inputPath: {self.inputPath}, ' \
            f'sheetName: {self.sheetName}, ' \
            f'outputPath: {self.outputPath}'


class ArgsError(Exception):
    """
    An error regarding the command line arguments.
    """


def excel2txt(workbook: Union[str, Workbook], outputPath: str = None) -> None:
    """
    Convert the entire excel file, including all its sheets, into one txt file.

    Parameters:
        workbook: The excel file to be converted. Can either be the path to the file
            or the actual Workbook object.
        outputPath: The path of the output txt file. If None, file will be placed
            in the same directory as the script.
    """
    # If the given parameter is a path, load the workbook
    if isinstance(workbook, str):
        workbook = loadExcel(workbook, readOnly=True, dataOnly=True)

    # If the output path is not provided, set the default
    if outputPath is None:
        outputPath = './workbook.output.txt'

    with open(outputPath, 'w', encoding='utf-8') as outputFile:
        for sheetName in sorted(workbook.sheetnames):
            outputFile.write(f'##########  {sheetName}  ##########\n\n')
            lines = _getText(workbook[sheetName])
            lines.append('')
            lines.append('')
            for line in lines:
                outputFile.write(line + '\n')


def sheet2txt(workbook: Union[str, Workbook], sheetName: str,
              outputPath: Optional[str] = None) -> None:
    """
    Convert a specified excel sheet into a txt file.

    Parameters:
        workbook: The excel file to be converted. Can either be the path to the file
            or the actual Workbook object.
        sheetName: The name of the sheet to be converted.
        outputPath: The path of the output txt file. If None, file will be placed
            in the same directory as the script.
    """
    # If the given parameter is a path, load the workbook
    if isinstance(workbook, str):
        workbook = loadExcel(workbook, readOnly=True, dataOnly=True)

    # If the output path is not provided, set the default
    if outputPath is None:
        outputPath = f'./{sanitizeFilename(sheetName)}.txt'

    with open(outputPath, 'w', encoding='utf-8') as outputFile:
        for line in _getText(workbook[sheetName]):
            outputFile.write(line + '\n')


def _getText(worksheet: Worksheet) -> List[str]:
    """
    Converts the worksheet into a list of strings.

    Parameters:
        worksheet: The worksheet to be converted.
    """
    result: List[str] = []
    for row in worksheet.iter_rows(values_only=True):
        line = ", ".join([str(value) for value in row])
        result.append(line)
    return result


def loadExcel(filePath: str, readOnly: bool = False, dataOnly: bool = False) -> Workbook:
    """
    Load the excel file. Returns a `Workbook` object.

    Parameters:
        filePath: The path of the excel file to be loaded.
        readOnly: If true, the file be loaded as a read-only workbook.
        dataOnly: If true, only the cell values will be retrieved.
    """
    return load_workbook(filePath, read_only=readOnly, data_only=dataOnly)


def sanitizeFilename(filename: str) -> str:
    """
    Remove invalid characters <>:"/\\|?* from the filename.
    """
    result = ''
    for c in filename:
        if c not in '<>:"/\\|?*':
            result += c
    return result


def _printUsage() -> None:
    """
    Print the usage instructions.
    """
    print('Usage:')
    print('   python -m excel2txt INPUT_FILEPATH [-s SHEET_NAME] [-o OUTPUT_FILEPATH]')


def _parseArgs(args: List[str]) -> Arguments:
    """
    Parse the arguments. Terminates the script if errors are found.
    """
    argLen = len(args)

    # Initialize the argument values
    inputPath: str = None
    sheetName: Optional[str] = None
    outputPath: Optional[str] = None

    # Check if the input path was specified
    if argLen < 2:
        raise ArgsError('The input file was not specified.')

    # Check if the input file exists
    if not os.path.exists(args[1]):
        raise ArgsError(f'The file "{args[1]}" does not exist.')

    inputPath = args[1]
    argIdx = 2

    # Check each optional argument
    while argIdx < argLen:

        # Check the sheet argument
        if args[argIdx] in ('-s', '--sheet'):
            if argIdx + 1 == argLen:
                raise ArgsError('Sheet name was not specified.')
            sheetName = args[argIdx + 1]
            argIdx += 2

        # Check the outputPath argument
        elif args[argIdx] in ('-o', '--output'):
            if argIdx + 1 == argLen:
                raise ArgsError('Output path was not specified.')
            outputPath = args[argIdx + 1]
            argIdx += 2

        # If the argument is unrecognized
        else:
            raise ArgsError(f'The argument "{args[2]}" is unrecognized.')

    return Arguments(inputPath,
                     sheetName,
                     outputPath)


if __name__ == '__main__':
    if '-h' in sys.argv or '--help' in sys.argv:
        _printUsage()
        sys.exit()

    try:
        _args = _parseArgs(sys.argv)
    except ArgsError as err:
        print(f'Error: {str(err)}')
        _printUsage()
        sys.exit(1)

    _workbook = loadExcel(_args.inputPath, readOnly=True, dataOnly=True)
    _sheetName = _args.sheetName
    _outputPath = _args.outputPath

    if _args.sheetName is not None:
        sheet2txt(_workbook, _sheetName, _outputPath)
    else:
        excel2txt(_workbook, _outputPath)
